"""
Excel / CSV 파일 파싱 서비스 (사양서 §4.2)

- .xlsx / .xls: openpyxl
- .csv:         csv (표준 라이브러리, pandas 없이도 동작)
- 반환: 시트 목록, 컬럼 목록, 10행 미리보기, 파싱 오류 행 목록
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field


@dataclass
class ParseResult:
    sheets: list[str]              # 시트 목록 (CSV면 ["Sheet1"])
    columns: list[str]             # 선택된 시트의 컬럼명 목록
    preview: list[list[str]]       # 최대 10행 미리보기 (문자열 그대로)
    total_rows: int                # 헤더 제외 전체 행 수
    parse_errors: list[ParseError] # 파싱 실패 행 정보


@dataclass
class ParseError:
    row: int           # 1-based 행 번호 (헤더 = 1)
    content: str       # 해당 행의 원본 내용 (요약)
    reason: str        # 실패 이유


@dataclass
class ExtractResult:
    data: list[float]               # 성공 파싱 수치값
    parse_errors: list[ParseError]  # 변환 실패 행
    skipped_count: int              # 빈 셀 / 문자열 스킵 수


# ── 메인 파싱 함수 ────────────────────────────────────────────────────────────

def parse_file(
    content: bytes,
    filename: str,
    detected_format: str,
    *,
    sheet_name: str | None = None,   # None이면 첫 번째 시트
    column_index: int | None = None, # None이면 첫 번째 숫자 열 자동 선택
    has_header: bool = True,
) -> ParseResult:
    """
    파일을 파싱해 시트 목록, 컬럼, 미리보기를 반환한다.

    Args:
        content:          파일 바이너리
        filename:         원본 파일명 (확장자 판별 보조)
        detected_format:  'xlsx' | 'xls' | 'csv'
        sheet_name:       특정 시트명 (None이면 첫 시트)
        column_index:     측정값 열 인덱스 (None이면 첫 번째 숫자 열)
        has_header:       첫 행을 헤더로 처리할지 여부

    Returns:
        ParseResult
    """
    if detected_format in ("xlsx", "xls"):
        return _parse_excel(content, sheet_name=sheet_name, has_header=has_header)
    elif detected_format == "csv":
        return _parse_csv(content, has_header=has_header)
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {detected_format}")


def extract_column(
    content: bytes,
    detected_format: str,
    *,
    sheet_name: str | None = None,
    column_index: int = 0,
    has_header: bool = True,
) -> ExtractResult:
    """
    특정 열의 수치 데이터를 추출한다.
    문자열·빈 셀은 자동 스킵, 변환 실패 행은 ParseError 목록으로 반환.
    """
    if detected_format in ("xlsx", "xls"):
        raw_rows = _read_excel_rows(content, sheet_name=sheet_name, has_header=has_header)
    else:
        raw_rows = _read_csv_rows(content, has_header=has_header)

    data: list[float] = []
    errors: list[ParseError] = []
    skipped = 0
    header_offset = 2 if has_header else 1  # 1-based 행 번호 기준

    for i, row in enumerate(raw_rows):
        row_num = i + header_offset
        if column_index >= len(row):
            skipped += 1
            continue

        cell = str(row[column_index]).strip()
        if not cell:
            skipped += 1
            continue

        try:
            data.append(float(cell))
        except ValueError:
            errors.append(ParseError(
                row=row_num,
                content=cell[:100],
                reason="숫자로 변환할 수 없습니다.",
            ))

    return ExtractResult(data=data, parse_errors=errors, skipped_count=skipped)


# ── Excel 파서 ────────────────────────────────────────────────────────────────

def _parse_excel(
    content: bytes,
    *,
    sheet_name: str | None,
    has_header: bool,
) -> ParseResult:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl이 설치되어 있지 않습니다.")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = wb.sheetnames

    ws = wb[sheet_name] if sheet_name and sheet_name in wb else wb.active
    active_sheet = ws.title

    rows_iter = ws.iter_rows(values_only=True)

    columns: list[str] = []
    preview: list[list[str]] = []
    parse_errors: list[ParseError] = []
    total_rows = 0

    all_rows: list[tuple] = list(rows_iter)

    if has_header and all_rows:
        header_row = all_rows[0]
        columns = [str(c) if c is not None else f"Column{i+1}"
                   for i, c in enumerate(header_row)]
        data_rows = all_rows[1:]
    else:
        data_rows = all_rows
        if data_rows:
            columns = [f"Column{i+1}" for i in range(len(data_rows[0]))]

    total_rows = len(data_rows)

    for i, row in enumerate(data_rows[:10]):
        preview.append([str(c) if c is not None else "" for c in row])

    wb.close()
    return ParseResult(
        sheets=[active_sheet],  # 선택된 시트만 반환 (목록은 별도)
        columns=columns,
        preview=preview,
        total_rows=total_rows,
        parse_errors=parse_errors,
    )


def _read_excel_rows(
    content: bytes,
    *,
    sheet_name: str | None,
    has_header: bool,
) -> list[tuple]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb else wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return all_rows[1:] if has_header and all_rows else all_rows


def _get_all_sheets(content: bytes) -> list[str]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


# ── CSV 파서 ─────────────────────────────────────────────────────────────────

def _parse_csv(content: bytes, *, has_header: bool) -> ParseResult:
    text = _decode_csv(content)
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)

    columns: list[str] = []
    parse_errors: list[ParseError] = []

    if has_header and all_rows:
        columns = [c.strip() for c in all_rows[0]]
        data_rows = all_rows[1:]
    else:
        data_rows = all_rows
        if data_rows:
            columns = [f"Column{i+1}" for i in range(len(data_rows[0]))]

    preview = [list(row) for row in data_rows[:10]]
    return ParseResult(
        sheets=["Sheet1"],
        columns=columns,
        preview=preview,
        total_rows=len(data_rows),
        parse_errors=parse_errors,
    )


def _read_csv_rows(content: bytes, *, has_header: bool) -> list[list[str]]:
    text = _decode_csv(content)
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    return all_rows[1:] if has_header and all_rows else all_rows


def _decode_csv(content: bytes) -> str:
    """UTF-8 → CP949(EUC-KR) 순으로 디코딩 시도 (한국어 파일 대응)."""
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return content.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return content.decode("latin-1")   # 최후 폴백
