"""
Excel 보고서 생성 서비스 (사양서 §11)

구성:
  Sheet 1 "보고서" — 분석 결과 + 데이터 + BarChart (단일 시트)
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from app.schemas.report import ReportRequest


# ── 스타일 팔레트 ─────────────────────────────────────────────────────────────
_BLUE_FILL  = PatternFill("solid", fgColor="1F4E79")
_GRAY_FILL  = PatternFill("solid", fgColor="D6E4F0")
_GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
_RED_FILL   = PatternFill("solid", fgColor="FCE4D6")
_WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
_BOLD_FONT  = Font(bold=True, size=10)
_NORM_FONT  = Font(size=10)
_CENTER     = Alignment(horizontal="center", vertical="center")
_LEFT       = Alignment(horizontal="left",   vertical="center")
_THIN_SIDE  = Side(style="thin", color="AAAAAA")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                      top=_THIN_SIDE, bottom=_THIN_SIDE)


def _header_cell(ws, row: int, col: int, value, width_hint: int = 20) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _BLUE_FILL
    cell.font      = _WHITE_FONT
    cell.alignment = _CENTER
    cell.border    = _THIN_BORDER
    ws.column_dimensions[get_column_letter(col)].width = width_hint


def _label_cell(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _GRAY_FILL
    cell.font      = _BOLD_FONT
    cell.alignment = _LEFT
    cell.border    = _THIN_BORDER


def _value_cell(ws, row: int, col: int, value, fmt: str = "") -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _NORM_FONT
    cell.alignment = _CENTER
    cell.border    = _THIN_BORDER
    if fmt:
        cell.number_format = fmt


def _grade_fill(cpk_val: float) -> PatternFill:
    if cpk_val >= 1.67:
        return PatternFill("solid", fgColor="70AD47")   # green
    if cpk_val >= 1.33:
        return PatternFill("solid", fgColor="FFD966")   # yellow
    return PatternFill("solid", fgColor="FF0000")        # red


def _cpk_grade(v: float) -> str:
    if v >= 2.0:  return "A++ (6σ)"
    if v >= 1.67: return "A+  (5σ)"
    if v >= 1.50: return "A   (≥5σ)"
    if v >= 1.33: return "B   (4σ)"
    if v >= 1.0:  return "C   (3σ)"
    return "D   (<3σ)"


# ── 데이터 열 (F-G) ────────────────────────────────────────────────────────────

def _build_data_columns(ws, req: "ReportRequest") -> tuple[int, int]:
    """
    F열(col 6): 순번, G열(col 7): 측정값
    헤더는 row 2부터 시작 (row 1은 제목 병합).
    데이터 마지막 행 번호를 반환.
    """
    DATA_COL_IDX  = 6   # F
    DATA_COL_VAL  = 7   # G
    HEADER_ROW    = 2

    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 18

    _header_cell(ws, HEADER_ROW, DATA_COL_IDX, "순번",   10)
    _header_cell(ws, HEADER_ROW, DATA_COL_VAL, "측정값", 18)

    for i, v in enumerate(req.data, start=1):
        row = HEADER_ROW + i
        _value_cell(ws, row, DATA_COL_IDX, i,   "#,##0")
        _value_cell(ws, row, DATA_COL_VAL, v, "#,##0.0000")

    last_data_row = HEADER_ROW + len(req.data)
    ws.auto_filter.ref = f"F{HEADER_ROW}:G{last_data_row}"
    return HEADER_ROW, last_data_row


# ── BarChart ──────────────────────────────────────────────────────────────────

def _build_chart(ws, header_row: int, last_data_row: int) -> None:
    """측정값 BarChart를 G열 데이터 기준으로 생성해 차트 아래에 삽입."""
    chart = BarChart()
    chart.type    = "col"
    chart.grouping = "clustered"
    chart.title   = "측정값 분포"
    chart.y_axis.title = "측정값"
    chart.x_axis.title = "순번"
    chart.style   = 10
    chart.width   = 20
    chart.height  = 14

    # 측정값 시리즈 (G열)
    data_ref = Reference(
        ws,
        min_col=7, max_col=7,
        min_row=header_row, max_row=last_data_row,
    )
    chart.add_data(data_ref, titles_from_data=True)

    # 순번 카테고리 (F열)
    cats = Reference(
        ws,
        min_col=6, max_col=6,
        min_row=header_row + 1, max_row=last_data_row,
    )
    chart.set_categories(cats)

    # 차트를 데이터 아래에 배치 (데이터 끝 + 2행)
    chart_anchor_row = last_data_row + 2
    ws.add_chart(chart, f"F{chart_anchor_row}")


# ── 분석 결과 열 (A-D) ────────────────────────────────────────────────────────

def _build_result_columns(ws, req: "ReportRequest") -> None:
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    r = 2  # row 1은 제목 병합

    # ── 메타 정보 ────────────────────────────────────────────────────────────
    for label, value in [
        ("분석 ID",   req.analysis_id),
        ("분석 모드",  req.mode.upper()),
        ("생성 일시",  generated_at),
    ]:
        _label_cell(ws, r, 1, label)
        cell = ws.cell(row=r, column=2, value=value)
        cell.font      = _NORM_FONT
        cell.alignment = _LEFT
        cell.border    = _THIN_BORDER
        ws.merge_cells(f"B{r}:D{r}")
        r += 1

    r += 1

    # ── 규격 ────────────────────────────────────────────────────────────────
    _header_cell(ws, r, 1, "규격 (Specification)")
    ws.merge_cells(f"A{r}:D{r}")
    r += 1

    for label, value in [
        ("USL (상한 규격)",   req.usl),
        ("LSL (하한 규격)",   req.lsl),
        ("Nominal (목표값)",  req.nominal if req.nominal is not None else "—"),
        ("서브그룹 크기",      req.subgroup_size if req.mode != "ppk" else "—"),
        ("σ 추정 방식",        req.sigma_method.upper() if req.mode != "ppk" else "—"),
    ]:
        _label_cell(ws, r, 1, label)
        _value_cell(ws, r, 2, value, "#,##0.0000" if isinstance(value, float) else "")
        ws.merge_cells(f"B{r}:D{r}")
        r += 1

    r += 1

    # ── 기술통계 ──────────────────────────────────────────────────────────────
    _header_cell(ws, r, 1, "기술통계 (Descriptive Statistics)")
    ws.merge_cells(f"A{r}:D{r}")
    r += 1

    s = req.stats
    for label, value in [
        ("데이터 수 (n)",             s.n),
        ("평균 (Mean)",               s.mean),
        ("전체 표준편차 (σ_overall)", s.std_overall),
        ("최솟값 (Min)",              s.min),
        ("최댓값 (Max)",              s.max),
        ("중위값 (Median)",           s.median),
    ]:
        _label_cell(ws, r, 1, label)
        fmt = "#,##0.0000" if isinstance(value, float) else "#,##0"
        _value_cell(ws, r, 2, value, fmt)
        ws.merge_cells(f"B{r}:D{r}")
        r += 1

    r += 1

    # ── Cpk 결과 ──────────────────────────────────────────────────────────────
    if req.cpk:
        c = req.cpk
        _header_cell(ws, r, 1, "Cpk 공정능력 지수")
        ws.merge_cells(f"A{r}:D{r}")
        r += 1

        for label, value in [
            ("Cpk",                   c.cpk),
            ("Cp",                    c.cp),
            ("Cpu (상측)",             c.cpu),
            ("Cpl (하측)",             c.cpl),
            ("σ̂ 추정 (σ_within)",     c.sigma_within),
            ("치우침 K",               c.k),
            ("불량률 ─ USL 초과 (%)", c.defect_usl_pct * 100),
            ("불량률 ─ LSL 미달 (%)", c.defect_lsl_pct * 100),
            ("불량률 합계 (%)",         c.defect_total_pct * 100),
            ("DPMO",                   c.dpmo),
            ("Sigma Level",            c.sigma_level),
        ]:
            _label_cell(ws, r, 1, label)
            cell = ws.cell(row=r, column=2, value=round(value, 6))
            cell.font          = _NORM_FONT
            cell.alignment     = _CENTER
            cell.border        = _THIN_BORDER
            cell.number_format = "#,##0.000000"
            ws.merge_cells(f"B{r}:D{r}")
            r += 1

        _label_cell(ws, r, 1, "Cpk 등급")
        grade_cell = ws.cell(row=r, column=2, value=_cpk_grade(c.cpk))
        grade_cell.font      = Font(bold=True, size=11)
        grade_cell.alignment = _CENTER
        grade_cell.fill      = _grade_fill(c.cpk)
        grade_cell.border    = _THIN_BORDER
        ws.merge_cells(f"B{r}:D{r}")
        r += 2

    # ── Ppk 결과 ──────────────────────────────────────────────────────────────
    if req.ppk:
        p = req.ppk
        _header_cell(ws, r, 1, "Ppk 공정성능 지수")
        ws.merge_cells(f"A{r}:D{r}")
        r += 1

        for label, value in [
            ("Ppk",                    p.ppk),
            ("Pp",                     p.pp),
            ("Ppu (상측)",              p.ppu),
            ("Ppl (하측)",              p.ppl),
            ("σ 전체 (σ_overall)",     p.sigma_overall),
            ("불량률 ─ USL 초과 (%)", p.defect_usl_pct * 100),
            ("불량률 ─ LSL 미달 (%)", p.defect_lsl_pct * 100),
            ("불량률 합계 (%)",         p.defect_total_pct * 100),
            ("DPMO",                   p.dpmo),
            ("Sigma Level",            p.sigma_level),
        ]:
            _label_cell(ws, r, 1, label)
            cell = ws.cell(row=r, column=2, value=round(value, 6))
            cell.font          = _NORM_FONT
            cell.alignment     = _CENTER
            cell.border        = _THIN_BORDER
            cell.number_format = "#,##0.000000"
            ws.merge_cells(f"B{r}:D{r}")
            r += 1

        _label_cell(ws, r, 1, "Ppk 등급")
        grade_cell = ws.cell(row=r, column=2, value=_cpk_grade(p.ppk))
        grade_cell.font      = Font(bold=True, size=11)
        grade_cell.alignment = _CENTER
        grade_cell.fill      = _grade_fill(p.ppk)
        grade_cell.border    = _THIN_BORDER
        ws.merge_cells(f"B{r}:D{r}")
        r += 2

    # ── 경고 ──────────────────────────────────────────────────────────────────
    if req.warnings:
        _header_cell(ws, r, 1, "경고 (Warnings)")
        ws.merge_cells(f"A{r}:D{r}")
        r += 1
        for w in req.warnings:
            cell = ws.cell(row=r, column=1, value=w)
            cell.fill      = PatternFill("solid", fgColor="FFF2CC")
            cell.font      = Font(size=9)
            cell.alignment = _LEFT
            cell.border    = _THIN_BORDER
            ws.merge_cells(f"A{r}:D{r}")
            r += 1


# ── 공개 API ─────────────────────────────────────────────────────────────────

def generate_excel(req: "ReportRequest") -> bytes:
    """
    ReportRequest 를 받아 xlsx 바이트를 반환한다.
    단일 시트 "보고서": 좌측(A-D) 분석 결과, 우측(F-G) 데이터 + BarChart.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "보고서"
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A2"

    # ── row 1: 전체 제목 ──────────────────────────────────────────────────────
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.fill      = _BLUE_FILL
        cell.font      = _WHITE_FONT
        cell.alignment = _CENTER
        cell.border    = _THIN_BORDER
    ws.cell(row=1, column=1, value="PCA 공정능력 분석 보고서")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 28

    # ── 좌측: 분석 결과 (A-D) ────────────────────────────────────────────────
    _build_result_columns(ws, req)

    # ── 우측: 데이터 목록 (F-G) ──────────────────────────────────────────────
    header_row, last_data_row = _build_data_columns(ws, req)

    # ── BarChart (데이터 아래) ────────────────────────────────────────────────
    _build_chart(ws, header_row, last_data_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
